# -*- coding: utf-8 -*-
"""
读取当前目录下 Vegetable 文件夹中的 Excel 表，
将表格第一行作为单位，读取商品名称和应发数量到该单位下，并打印；
并生成 蔬心兰.xlsx：以单位为列、商品为行的汇总表。

跨平台：可在 Windows、macOS、Linux 上运行。
"""

import os
import sys
import xlrd

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None


def _ensure_console_utf8():
    """在 Windows 下尽量使控制台输出 UTF-8，避免中文乱码。"""
    if sys.platform != "win32":
        return
    try:
        if hasattr(sys.stdout, "reconfigure"):
            sys.stdout.reconfigure(encoding="utf-8")
        if hasattr(sys.stderr, "reconfigure"):
            sys.stderr.reconfigure(encoding="utf-8")
    except (AttributeError, OSError):
        pass


def get_vegetable_dir():
    """Vegetable 目录路径（当前脚本所在目录下的 Vegetable 文件夹）。"""
    # 处理 EXE 打包后的情况
    if getattr(sys, 'frozen', False):
        # 获取 EXE 文件所在目录
        base = os.path.dirname(sys.executable)
    else:
        # 正常脚本运行情况
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "Vegetable")


def find_header_row(sheet):
    """找到表头行，返回 (行索引, 序号列索引, 商品名称列索引, 应发列索引)。"""
    for r in range(min(sheet.nrows, 15)):
        row = [str(sheet.cell_value(r, c)).strip() for c in range(sheet.ncols)]
        try:
            idx_name = row.index("商品名称")
        except ValueError:
            continue
        idx_yifa = None
        for i, cell in enumerate(row):
            if cell in ("应发", "应发数量"):
                idx_yifa = i
                break
        if idx_yifa is None:
            continue
        try:
            idx_xuhao = row.index("序号")
        except ValueError:
            idx_xuhao = None
        return r, idx_xuhao, idx_name, idx_yifa
    return None, None, None, None


def read_sheet(sheet):
    """从工作表中解析：单位（第一行）、商品序号、商品名称、应发数量。"""
    if sheet.nrows == 0:
        return None, []

    # 第一行作为单位（取第一个非空单元格或整行拼接）
    unit_cells = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
    unit = "".join(unit_cells).strip() or "(未填写单位)"

    header_row, col_xuhao, col_name, col_yifa = find_header_row(sheet)
    if header_row is None or col_name is None or col_yifa is None:
        return unit, []

    items = []
    for r in range(header_row + 1, sheet.nrows):
        name = str(sheet.cell_value(r, col_name)).strip()
        if not name:
            continue
        xuhao = ""
        if col_xuhao is not None:
            raw = sheet.cell_value(r, col_xuhao)
            if isinstance(raw, float) and raw == int(raw):
                xuhao = str(int(raw))
            else:
                xuhao = str(raw).strip() if raw else ""
        try:
            yifa = sheet.cell_value(r, col_yifa)
            if isinstance(yifa, (int, float)):
                yifa_num = int(yifa) if isinstance(yifa, float) and yifa == int(yifa) else yifa
            else:
                yifa_num = 0
        except (IndexError, TypeError):
            yifa_num = 0
        # 只保留应发数量 > 0 的商品行，过滤表尾汇总等
        if yifa_num > 0:
            items.append((xuhao, name, yifa_num))

    return unit, items


def read_excel(path):
    """读取一个 Excel 文件，支持 .xls。返回 [(单位, [(商品序号, 商品名称, 应发数量), ...]), ...]。"""
    results = []
    try:
        wb = xlrd.open_workbook(path)
    except Exception as e:
        print(f"  无法打开: {e}")
        return results
    for sheet in wb.sheets():
        unit, items = read_sheet(sheet)
        if unit or items:
            results.append((unit, items))
    return results


def collect_all_data(veg_dir, excel_files):
    """从所有 Excel 文件中收集 (单位, [(商品, 应发数量), ...])。"""
    all_data = []
    for filename in excel_files:
        path = os.path.join(veg_dir, filename)
        for unit, items in read_excel(path):
            if items:
                all_data.append((unit, items))
    return all_data


def build_pivot_table(all_data):
    """
    按商品序号合并：同一序号的商品为一行，各单位数量相加。
    返回 (序号列表, {序号: 商品名称拼接}, 单位列表, {(序号, 单位): 数量})。
    """
    units_set = set()
    serial_to_names = {}  # 序号 -> set of 商品名称（合并后用 " / " 拼接）
    pivot = {}  # (序号, unit) -> qty（同序号同单位数量相加）
    for unit, items in all_data:
        units_set.add(unit)
        for xuhao, name, qty in items:
            if xuhao not in serial_to_names:
                serial_to_names[xuhao] = set()
            serial_to_names[xuhao].add(name)
            key = (xuhao, unit)
            pivot[key] = pivot.get(key, 0) + qty
    # 序号排序：有号的在前，空序号在后
    serials = sorted(serial_to_names.keys(), key=lambda x: (x == "", x))
    units = sorted(units_set)
    # 每个序号对应的商品名称（多个用 " / " 连接）
    serial_names = {xuhao: " / ".join(sorted(names)) for xuhao, names in serial_to_names.items()}
    return serials, serial_names, units, pivot


def write_shuxinlan_excel(serials, serial_names, units, pivot, output_path):
    """生成 蔬心兰.xlsx：按商品序号合并行，第一列序号，第二列商品名称，后续列为各单位。"""
    if Workbook is None:
        print("错误: 未安装 openpyxl，无法生成蔬心兰.xlsx", file=sys.stderr)
        print("请运行: pip install openpyxl", file=sys.stderr)
        sys.exit(1)
    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"
    ws.cell(row=1, column=1, value="序号")
    ws.cell(row=1, column=2, value="商品名称")
    for c, unit in enumerate(units, start=3):
        ws.cell(row=1, column=c, value=unit)
    for r, xuhao in enumerate(serials, start=2):
        ws.cell(row=r, column=1, value=xuhao)
        ws.cell(row=r, column=2, value=serial_names.get(xuhao, ""))
        for c, unit in enumerate(units, start=3):
            val = pivot.get((xuhao, unit), "")
            ws.cell(row=r, column=c, value=val if val != "" else "")
    wb.save(output_path)


def main():
    _ensure_console_utf8()
    veg_dir = get_vegetable_dir()
    print(f"路径：{veg_dir}")

    # 如果是 EXE 文件且目录不存在，显示友好提示
    is_exe = getattr(sys, 'frozen', False)
    if is_exe and not os.path.isdir(veg_dir):
        print("错误: 找不到 Vegetable 文件夹")
        print("请在程序所在目录创建 Vegetable 文件夹")
        print("并将 Excel 文件放入其中")
        input("按回车键退出...")
        return

    if not os.path.isdir(veg_dir):
        print(f"目录不存在: {veg_dir}")
        return

    excel_files = [
        f for f in os.listdir(veg_dir)
        if f.endswith((".xls", ".xlsx")) and not f.startswith("~")
    ]
    excel_files.sort()

    if not excel_files:
        print(f"在 {veg_dir} 下未找到 Excel 文件。")
        return

    print(f"共找到 {len(excel_files)} 个 Excel 文件\n")
    print("=" * 60)

    for filename in excel_files:
        path = os.path.join(veg_dir, filename)
        print(f"\n文件: {filename}")
        for unit, items in read_excel(path):
            print(f"  单位: {unit}")
            if not items:
                print("    （无商品数据）")
            else:
                for xuhao, name, yifa in items:
                    print(f"    - {name}  应发数量: {yifa}")
        print("-" * 60)

    # 生成蔬心兰.xlsx：单位为列，商品为行
    all_data = collect_all_data(veg_dir, excel_files)
    if all_data:
        serials, serial_names, units, pivot = build_pivot_table(all_data)
        # 使用与 Vegetable 文件夹相同的基础目录
        if getattr(sys, 'frozen', False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
        out_path = os.path.join(base_dir, "蔬心兰.xlsx")
        write_shuxinlan_excel(serials, serial_names, units, pivot, out_path)
        print(f"\n已生成汇总表: {out_path}")
        print(f"  行（按商品序号合并）: {len(serials)}，列（单位）: {len(units)}")
    else:
        print("\n无数据，未生成蔬心兰.xlsx")


if __name__ == "__main__":
    main()
